### Spec ###
# LinePay 匯款明細整理程式
# 用途：將 linepay明細.xlsx 的交易資料按「撥款預定日」分組整理
# 主要功能：
#   1. 掃描 報表/ 底下所有月份資料夾（YYYYMM），自動判斷哪些尚未處理
#   2. 讀取資料夾內的 linepay明細.xlsx
#   3. 以撥款預定日為大分類，每個撥款預定日底下按交易日分組
#   4. 加總每個交易日的付款金額、手續費合計、排定的各項目撥款（實收）
#   5. 輸出格式化 Excel，含合併儲存格、粗體、千分位、小計與總計
# 輸入：報表/{YYYYMM}/linepay明細.xlsx
# 輸出：報表/{YYYYMM}/LinePay匯款明細整理_YYYYMM.xlsx
# 關聯：參考電子支付對帳程式的 base_dir 定位方式與 openpyxl 格式模式
### End Spec ###

import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def fmt_date(d):
    """將 20260128 格式轉為 2026/01/28"""
    s = str(int(d))
    return f"{s[:4]}/{s[4:6]}/{s[6:8]}"


def process_folder(folder_path: str, folder_name: str):
    """處理單一月份資料夾的 LinePay 匯款明細整理"""
    # 檢查輸出檔是否已存在
    output_name = f"LinePay匯款明細整理_{folder_name}.xlsx"
    output_path = os.path.join(folder_path, output_name)
    if os.path.exists(output_path):
        print(f"  已有 {output_name}，跳過")
        return

    # 檢查輸入檔
    filepath = os.path.join(folder_path, 'linepay明細.xlsx')
    if not os.path.exists(filepath):
        print(f"  找不到 linepay明細.xlsx，跳過")
        return

    # 讀取資料
    df = pd.read_excel(filepath)
    print(f"  讀取 {len(df)} 筆交易")

    # 分組：撥款預定日 -> 交易日 -> 加總
    grouped = df.groupby(['撥款預定日', '交易日']).agg(
        付款金額=('付款金額', 'sum'),
        手續費=('手續費合計', 'sum'),
        實收=('排定的各項目撥款', 'sum'),
    ).reset_index().sort_values(['撥款預定日', '交易日'])

    # 建立 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '匯款明細整理'

    # 樣式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    group_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    group_font = Font(bold=True, size=11)
    subtotal_font = Font(bold=True)
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # 欄寬
    col_widths = [16, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    grand_payment = 0
    grand_fee = 0
    grand_received = 0

    for payout_date, group in grouped.groupby('撥款預定日'):
        # 撥款預定日標題列
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=f"撥款預定日: {fmt_date(payout_date)}")
        cell.font = group_font
        cell.fill = group_fill
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = group_fill
        row += 1

        # 欄位標題
        headers = ['交易日', '付款金額', '手續費', '實收', '手續費比例']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # 資料列
        sub_payment = 0
        sub_fee = 0
        sub_received = 0

        for _, r in group.iterrows():
            ws.cell(row=row, column=1, value=fmt_date(r['交易日']))
            ws.cell(row=row, column=2, value=int(r['付款金額']))
            ws.cell(row=row, column=3, value=round(r['手續費'], 2))
            ws.cell(row=row, column=4, value=round(r['實收'], 2))

            # 手續費比例
            payment = int(r['付款金額'])
            ratio = r['手續費'] / payment if payment else 0
            ws.cell(row=row, column=5, value=ratio)
            ws.cell(row=row, column=5).number_format = '0.00%'

            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = thin_border

            sub_payment += int(r['付款金額'])
            sub_fee += r['手續費']
            sub_received += r['實收']
            row += 1

        # 小計列
        ws.cell(row=row, column=1, value='小計')
        ws.cell(row=row, column=2, value=sub_payment)
        ws.cell(row=row, column=3, value=round(sub_fee, 2))
        ws.cell(row=row, column=4, value=round(sub_received, 2))
        sub_ratio = sub_fee / sub_payment if sub_payment else 0
        ws.cell(row=row, column=5, value=sub_ratio)
        ws.cell(row=row, column=5).number_format = '0.00%'
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        for c in range(1, 6):
            ws.cell(row=row, column=c).font = subtotal_font
        row += 1

        # 空行
        row += 1

        grand_payment += sub_payment
        grand_fee += sub_fee
        grand_received += sub_received

    # 總計列
    ws.cell(row=row, column=1, value='總計')
    ws.cell(row=row, column=2, value=grand_payment)
    ws.cell(row=row, column=3, value=round(grand_fee, 2))
    ws.cell(row=row, column=4, value=round(grand_received, 2))
    grand_ratio = grand_fee / grand_payment if grand_payment else 0
    ws.cell(row=row, column=5, value=grand_ratio)
    ws.cell(row=row, column=5).number_format = '0.00%'
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    for c in range(1, 6):
        ws.cell(row=row, column=c).font = total_font
        ws.cell(row=row, column=c).fill = total_fill

    wb.save(output_path)

    print(f"  完成！已輸出：{output_name}")
    print(f"  共 {len(grouped['撥款預定日'].unique())} 個撥款預定日，{len(df)} 筆交易")
    print(f"  總計：付款金額 {grand_payment:,}，手續費 {grand_fee:,.2f}，實收 {grand_received:,.2f}")


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    report_dir = os.path.join(base_dir, '報表')

    if not os.path.isdir(report_dir):
        print("錯誤：找不到「報表」資料夾")
        input("按 Enter 結束...")
        return

    # 掃描所有月份資料夾
    folders = sorted([
        d for d in os.listdir(report_dir)
        if os.path.isdir(os.path.join(report_dir, d)) and re.match(r'^\d{6}$', d)
    ])

    if not folders:
        print("報表/ 底下沒有月份資料夾（格式：YYYYMM）")
        input("按 Enter 結束...")
        return

    print(f"找到 {len(folders)} 個月份資料夾：{', '.join(folders)}\n")

    for folder_name in folders:
        folder_path = os.path.join(report_dir, folder_name)
        print(f"[{folder_name}]")
        process_folder(folder_path, folder_name)
        print()

    input("按 Enter 結束...")


if __name__ == '__main__':
    main()
