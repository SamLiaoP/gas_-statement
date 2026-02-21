### Spec ###
# 加油站支付對帳系統
# 用途：自動比對內帳（115.XX.xlsx）與各支付管道（LinePay、中油Pay）的交易明細
# 主要功能：
#   1. 掃描 reports/ 底下所有月份資料夾（YYYYMM），自動判斷哪些尚未處理
#   2. 自動偵測資料夾內的內帳檔案（NNN.MM.xlsx 格式）
#   3. 讀取內帳金額表 sheet，解析各支付管道每日金額
#   4. 讀取各支付管道明細檔，解析每日交易金額
#   5. 比對內帳 vs 明細，產出對帳結果 Excel（含差異標記和紅色高亮）
# 輸入：reports/{YYYYMM}/115.XX.xlsx（內帳）、linepay明細.xlsx、中油pay明細.xls
# 輸出：reports/{YYYYMM}/對帳結果_YYYYMM.xlsx
# 執行方式：按兩下 main.py 或 python3 main.py
### End Spec ###

import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Callable, Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill

# --- 型別定義 ---
DayAmount = Dict[int, int]  # {日: 金額}


# --- Channel 設定 ---
@dataclass
class ChannelConfig:
    name: str           # 顯示名稱
    label: str          # 內帳 col 0 的列標籤
    detail_file: str    # 明細檔名
    parser: Callable    # (filepath, year, month) -> DayAmount


def parse_linepay(filepath: str, year: int, month: int) -> DayAmount:
    """解析 LinePay 明細，21:00 後交易算隔天"""
    df = pd.read_excel(filepath)
    result: DayAmount = {}
    for _, row in df.iterrows():
        ts = str(int(row['交易日期']))
        dt = datetime(int(ts[:4]), int(ts[4:6]), int(ts[6:8]),
                      int(ts[8:10]), int(ts[10:12]))
        # 21:00 規則：交易時間 >= 21:00 歸入隔天
        if dt.hour >= 21:
            dt += timedelta(days=1)
        if dt.year == year and dt.month == month:
            day = dt.day
            result[day] = result.get(day, 0) + int(row['付款金額'])
    return result


def parse_cpc(filepath: str, year: int, month: int) -> DayAmount:
    """解析中油Pay明細"""
    df = pd.read_excel(filepath, header=3)
    result: DayAmount = {}
    for _, row in df.iterrows():
        parts = str(row['交易日']).split('/')
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if y == year and m == month:
            amount = int(str(row['交易金額']).replace(',', ''))
            result[d] = result.get(d, 0) + amount
    return result


CHANNELS = [
    ChannelConfig("LINE PAY", "LINE PAY", "linepay明細.xlsx", parse_linepay),
    ChannelConfig("中油PAY", "中油PAY(CPC)", "中油pay明細.xls", parse_cpc),
]


def read_internal(filepath: str, label: str) -> DayAmount:
    """讀取內帳金額表，找到指定標籤的 row，回傳每日金額"""
    df = pd.read_excel(filepath, sheet_name='金額表', header=None)
    # 找目標 row
    target_row = None
    for i in range(df.shape[0]):
        if str(df.iloc[i, 0]).strip() == label:
            target_row = i
            break
    if target_row is None:
        print(f"錯誤：內帳中找不到「{label}」")
        return {}

    result: DayAmount = {}
    for day in range(1, 32):
        a_col = (day - 1) * 2 + 1
        b_col = (day - 1) * 2 + 2
        if b_col >= df.shape[1]:
            break
        a = pd.to_numeric(df.iloc[target_row, a_col], errors='coerce')
        b = pd.to_numeric(df.iloc[target_row, b_col], errors='coerce')
        total = int((0 if pd.isna(a) else a) + (0 if pd.isna(b) else b))
        if total != 0:
            result[day] = total
    return result


def find_internal_file(base_dir: str) -> tuple:
    """自動找內帳檔案，回傳 (filepath, 西元年, 月份) 或 None"""
    pattern = re.compile(r'^(\d+)\.(\d{2})\.xlsx$')
    matches = []
    for f in os.listdir(base_dir):
        m = pattern.match(f)
        if m and not f.startswith('~$'):
            roc_year = int(m.group(1))
            month = int(m.group(2))
            matches.append((f, roc_year + 1911, month))

    if not matches:
        return None

    if len(matches) == 1:
        f, year, month = matches[0]
        return os.path.join(base_dir, f), year, month

    # 多個檔案時取第一個（資料夾內理論上只有一個月份）
    f, year, month = matches[0]
    return os.path.join(base_dir, f), year, month


def compare_all(channels_data: list, year: int, month: int) -> pd.DataFrame:
    """
    比對所有管道，產出合併 DataFrame
    channels_data: [(name, internal_data, detail_data), ...]
    """
    all_days = set()
    for name, internal, detail in channels_data:
        all_days.update(internal.keys())
        all_days.update(detail.keys())
    all_days = sorted(all_days)

    rows = []
    for day in all_days:
        row = {'日期': day}
        for name, internal, detail in channels_data:
            iv = internal.get(day, 0)
            dv = detail.get(day, 0)
            diff = iv - dv
            row[f'{name} 內帳'] = iv
            row[f'{name} 明細'] = dv
            row[f'{name} 差異'] = diff
            row[f'{name} 狀態'] = 'V' if diff == 0 else 'X'
        rows.append(row)

    # 合計列
    total_row = {'日期': '合計'}
    for name, internal, detail in channels_data:
        total_i = sum(internal.values())
        total_d = sum(detail.values())
        total_row[f'{name} 內帳'] = total_i
        total_row[f'{name} 明細'] = total_d
        total_row[f'{name} 差異'] = total_i - total_d
        total_row[f'{name} 狀態'] = ''
    rows.append(total_row)

    return pd.DataFrame(rows)


def style_output(filepath: str, df: pd.DataFrame):
    """對輸出 Excel 加上格式：X 狀態欄標紅"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    ws = wb.active

    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red_font = Font(color='CC0000')

    status_cols = []
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name.endswith('狀態'):
            status_cols.append(col_idx)

    for row_idx in range(2, ws.max_row + 1):
        for sc in status_cols:
            if ws.cell(row=row_idx, column=sc).value == 'X':
                ws.cell(row=row_idx, column=sc).fill = red_fill
                ws.cell(row=row_idx, column=sc).font = red_font

    for col_idx, col_name in enumerate(df.columns, 1):
        if any(col_name.endswith(s) for s in ['內帳', '明細', '差異']):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0'

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ''
            width = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filepath)


def process_folder(folder_path: str, folder_name: str):
    """處理單一月份資料夾的對帳"""
    output_name = f"對帳結果_{folder_name}.xlsx"
    output_path = os.path.join(folder_path, output_name)
    if os.path.exists(output_path):
        print(f"  已有 {output_name}，跳過")
        return

    result = find_internal_file(folder_path)
    if result is None:
        print(f"  找不到內帳檔案，跳過")
        return
    internal_path, year, month = result
    print(f"  內帳：{os.path.basename(internal_path)}（{year}年{month}月）")

    channels_data = []
    for ch in CHANNELS:
        internal = read_internal(internal_path, ch.label)
        if not internal:
            print(f"  警告：內帳中 {ch.label} 無資料")

        detail_path = os.path.join(folder_path, ch.detail_file)
        if not os.path.exists(detail_path):
            print(f"  警告：找不到 {ch.detail_file}，跳過此管道")
            continue
        detail = ch.parser(detail_path, year, month)
        print(f"  {ch.name}：內帳 {len(internal)} 天，明細 {len(detail)} 天")

        channels_data.append((ch.name, internal, detail))

    if not channels_data:
        print(f"  沒有可比對的資料，跳過")
        return

    df = compare_all(channels_data, year, month)

    df.to_excel(output_path, index=False, sheet_name='對帳結果')
    style_output(output_path, df)
    print(f"  完成！已輸出：{output_name}")


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    report_dir = os.path.join(base_dir, 'reports')

    if not os.path.isdir(report_dir):
        print("錯誤：找不到「reports」資料夾")
        input("按 Enter 結束...")
        return

    folders = sorted([
        d for d in os.listdir(report_dir)
        if os.path.isdir(os.path.join(report_dir, d)) and re.match(r'^\d{6}$', d)
    ])

    if not folders:
        print("reports/ 底下沒有月份資料夾（格式：YYYYMM）")
        input("按 Enter 結束...")
        return

    print(f"找到 {len(folders)} 個月份資料夾：{', '.join(folders)}\n")

    for folder_name in folders:
        folder_path = os.path.join(report_dir, folder_name)
        print(f"[{folder_name}]")
        process_folder(folder_path, folder_name)
        print()

    input("按 Enter 結束...")


if __name__ == '__main__':
    main()
